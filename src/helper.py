from src.general_class import TableManger
import openpyxl
import os
import win32com.client as win32
import pythoncom
import time
from  tkinter import messagebox, ttk, simpledialog
import subprocess
import tkinter as tk
from tkinter.scrolledtext import ScrolledText
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import xlwings as xw

def get_item_pks(rfq_pk):
    """Gets all the ItemPK that is there in the RFQ"""
    rfq_line_table = TableManger("RequestForQuoteLine")
    quote_assembly_table = TableManger("QuoteAssembly")
    
    # Retrieve the list of (QuoteFK, Quantity) for the given RFQ
    quote_fk_list = rfq_line_table.get("QuoteFK", "Quantity", RequestForQuoteFK=rfq_pk)
    all_item_pks = []
    item_fks_dict = {}
    
    if quote_fk_list:
        for fk, quantity in quote_fk_list:
            if quantity is not None:
                # Retrieve the list of (ItemFK, QuantityRequired) for each QuoteFK
                item_pks = quote_assembly_table.get("ItemFK", "QuantityRequired", QuoteFK=fk)
                
                if item_pks:
                    for item_fk, qty_required in item_pks:
                        if qty_required is not None:
                            # Calculate total quantity required for each item
                            total_quantity = qty_required * quantity
                            all_item_pks.append((item_fk, total_quantity))
                            
        # Aggregate quantities for each item
        for item_fk, quantity in all_item_pks:
            if item_fk is not None:
                if item_fk in item_fks_dict:
                    item_fks_dict[item_fk] += quantity
                else:
                    item_fks_dict[item_fk] = quantity
    
    return item_fks_dict

def get_items_dict():
    item_table = TableManger("Item")
    my_dict = {}
    data = item_table.get("ItemPK", "PartNumber")
    for a in data: 
        my_dict[a[0]] = a[1]
    return my_dict

def get_rfq_pk():
    rfq_table = TableManger("RequestForQuote")
    return rfq_table.get("RequestForQuotePK")

def create_single_item_dict(item_id, qty_req):
    """Creates a dictionary for a single item with its required quantity as values"""
    item_dict = {}
    item_dict[item_id] = qty_req
    return item_dict

def get_item_dict(item_fks_dict):
    """"Creates a dictionary with all the details for an item as values and its PK as key"""
    item_table = TableManger("Item")
    item_dict = {}
    
    for fk, value in item_fks_dict.items():
        details = item_table.get(
            "PartNumber", "Description", "ItemTypeFK", "PartLength",
            "PartWidth", "Thickness", "StockLength", "StockWidth", "PurchaseOrderComment",
            ItemPK=fk
        )
        if details:
            columns = ["PartNumber", "Description", "ItemTypeFK", "PartLength",
                       "PartWidth", "Thickness", "StockLength", "StockWidth", "PurchaseOrderComment"]
            item_details = details[0]
            item_dict[fk] = {column: value for column, value in zip(columns, item_details)}
            item_dict[fk]["QuantityRequired"] = value
    
    return item_dict

def get_email_groups():
    """ Returns all emails that belong to the following groups
        'material', 'hardware', 'op', 'tooling'. 
        {'material': <values>, 'op': <values> ...}
     """
    party_buyer_table = TableManger("PartyBuyer") 
    party_table = TableManger("Party")
    parties = {                              # name : partyFK (we get buyerfk from this)
        "fin" : 3755,
        "mat-al": 3744,
        "mat-steel": 3747,
        "ht": 3764,
        'hardware': 3867,
    }
    email_dict = {}
    for key, partyfk in parties.items():
        buyerfk = party_buyer_table.get("BuyerFK", PartyFK=partyfk)
        email_dict[key] = [] 
        for fk in buyerfk:
            email = party_table.get("Email", PartyPK=fk[0])[0][0]
            email_dict[key].append(email)
 
    return email_dict
 
 
# the categories in the below function are meant to be directly used in the email_dict returned in get_email_groups
# the category that is appended to the item dict will match the key in email dict returned in get_email_groups
 

def sort_items_in_groups(item_dict: dict):
    """Identifies item type and then appends the category as a key, value pair to the item_dict."""
    categories = ["standard", "mat", "hardware", "msc", "fin", "kit", "tooling"]
    
    for itempk, column_values in item_dict.items():
        item_type_index = item_dict[itempk]["ItemTypeFK"] - 1  # to get the corresponding index in categories
        category = categories[item_type_index]
        
        part_number_split = item_dict[itempk]["PartNumber"].lower().split()
        email_category = None  # Initialize email_category to avoid reference before assignment

        if category == "fin":
            if "ht" in part_number_split or "heat" in part_number_split:
                email_category = "ht"
            else:
                email_category = "fin"
        elif category in ["hardware", "tooling"]:
            email_category = "hardware"
        elif category == "mat":
            if "al" in part_number_split:
                email_category = "mat-al"
            elif "steel" in part_number_split or "st" in part_number_split:
                email_category = "mat-steel"
            else:
                email_category = "mat-al"
        else:
            email_category = "mat-al"  # Default email_category to the category if not specified
        
        item_dict[itempk]["Category"] = category
        item_dict[itempk]["EmailCategory"] = email_category
 
    return item_dict

def create_excel(filepath, item_dict, rfq_number, item_id):
    """Creates an Excel file with the item_dict data for a RFQ Number."""
    workbook = openpyxl.load_workbook(filepath)
    if rfq_number:
        new_dir = os.path.abspath(os.path.join(".", "RFQ_Excel", str(rfq_number)))
        filename_parts = filepath.split("_")[2:]
        new_filename = f"RFQ_{rfq_number}_" + "_".join(filename_parts)
        new_filepath = os.path.join(new_dir, new_filename)
    elif item_id:
        new_dir = os.path.abspath(os.path.join(".", "RFQ_Excel", str(item_id)))
        filename_parts = filepath.split("_")[2:]
        new_filename = f"RFQ_{item_id}_" + "_".join(filename_parts)
        new_filepath = os.path.join(new_dir, new_filename)

    os.makedirs(new_dir, exist_ok=True)
    sheet = workbook.active
    sheet.delete_rows(7, sheet.max_row - 6)
    row_idx = 7

    for key, values in item_dict.items():
        if filepath == f"RFQ_template_{values['EmailCategory']}.xlsx":
            sheet.cell(row=row_idx, column=1).value = values['PartNumber']
            sheet.cell(row=row_idx, column=2).value = values['Description']
            sheet.cell(row=row_idx, column=6).value = values['StockLength'] if values['EmailCategory'] in ['mat-al', 'mat-steel'] else values['PartLength']
            sheet.cell(row=row_idx, column=5).value = values['StockWidth'] if values['EmailCategory'] == 'mat-al' or values['Category'] == 'mat-steel' else values['PartWidth']
            sheet.cell(row=row_idx, column=4).value = values['Thickness']
            sheet.cell(row=row_idx, column=7).value = values['PurchaseOrderComment']
            sheet.cell(row=row_idx, column=3).value = values['QuantityRequired']

            # Apply wrap text to relevant cells
            for col in range(1, 8):
                sheet.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True)
            
            row_idx += 1

    workbook.save(new_filepath)
    auto_fit_excel(new_filepath)
    return new_filepath

def auto_fit_excel(file_path):
    # Open the Excel workbook
    app = xw.App(visible=False)
    workbook = xw.Book(file_path)
 
    # Loop through each sheet in the workbook
    for sheet in workbook.sheets:
        # Auto-fit columns
        sheet.autofit('c')
        # Auto-fit rows
        sheet.autofit('r')
 
    # Save and close the workbook
    workbook.save(file_path)
    workbook.close()
    app.quit()

# def create_excel(filepath, item_dict, rfq_number, item_id):
#     """Creates an Excel file with the item_dict data for a RFQ Number"""
#     workbook = openpyxl.load_workbook(filepath)
#     if rfq_number:
#         new_dir = os.path.abspath(os.path.join(".", "RFQ_Excel", str(rfq_number)))
#         filename_parts = filepath.split("_")[2:]
#         new_filename = f"RFQ_{rfq_number}_" + "_".join(filename_parts)
#         new_filepath = os.path.join(new_dir, new_filename)
#     elif item_id:
#         new_dir = os.path.abspath(os.path.join(".", "RFQ_Excel", str(item_id)))
#         filename_parts = filepath.split("_")[2:]
#         new_filename = f"RFQ_{item_id}_" + "_".join(filename_parts)
#         new_filepath = os.path.join(new_dir, new_filename)

#     os.makedirs(new_dir, exist_ok=True)
#     sheet = workbook.active
#     sheet.delete_rows(3, sheet.max_row - 2)
#     row_idx = 3

#     for key, values in item_dict.items():
#         if filepath == f"RFQ_template_{values['EmailCategory']}.xlsx":
#             sheet.cell(row=row_idx, column=1).value = values['PartNumber']
#             sheet.cell(row=row_idx, column=2).value = values['Description']
#             sheet.cell(row=row_idx, column=6).value = values['StockLength'] if values['EmailCategory'] in ['mat-al', 'mat-steel'] else values['PartLength']
#             sheet.cell(row=row_idx, column=5).value = values['StockWidth'] if values['EmailCategory'] == 'mat-al' or values['Category'] == 'mat-steel' else values['PartWidth']
#             sheet.cell(row=row_idx, column=4).value = values['Thickness']
#             sheet.cell(row=row_idx, column=7).value = values['PurchaseOrderComment']
#             sheet.cell(row=row_idx, column=3).value = values['QuantityRequired']

#             # Apply wrap text to relevant cells, especially the "Comment" column (column 7)
#             for col in range(1, 8):
#                 sheet.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True)
            
#             row_idx += 1

#     # Adjust column widths based on the length of the text in each cell, including the header
#     max_comment_width = 50  # Define a reasonable max width for the "Comment" column
#     for column_cells in sheet.columns:
#         max_length = 0
#         column = column_cells[0].column_letter  # Get the column letter
#         for cell in column_cells:
#             try:
#                 if cell.value and len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
#             except:
#                 pass
        
#         # Adjust column width with consideration for "Comment" column max width
#         adjusted_width = (max_length + 2)
#         if column == "G":  # Column "G" corresponds to the "Comment" column
#             adjusted_width = min(adjusted_width, max_comment_width)
#         sheet.column_dimensions[column].width = adjusted_width

#     # Adjust row heights based on the length of the text in each cell, including the heading row
#     for row in sheet.iter_rows(min_row=1, max_row=row_idx - 1):
#         max_height = 15  # Default row height
#         for cell in row:
#             if cell.value and isinstance(cell.value, str):
#                 lines = cell.value.split('\n')
#                 height = len(lines) * 15  # Approximate row height for multiple lines of text
#                 if height > max_height:
#                     max_height = height
#         sheet.row_dimensions[row[0].row].height = max_height

#     # Explicitly adjust the width of columns based on the header text, adding extra width
#     extra_width = 5  # Define the extra width for the headers
#     for cell in sheet[1]:
#         if cell.value:
#             column = get_column_letter(cell.column)
#             adjusted_width = len(str(cell.value)) + 2 + extra_width
#             sheet.column_dimensions[column].width = max(sheet.column_dimensions[column].width, adjusted_width)

#     workbook.save(new_filepath)
#     return new_filepath


def send_all_emails(filepath_folder_of_excel_sheets: str) -> None:
    """Sends all emails to the IDs listed in the email sheet for each category"""
    pass


def create_excel_sheets(rfq_number=None, item_id=None, qty_req=None):
    """Creates Excel sheet and returns the filepath of the excel sheet"""
    if rfq_number:
        item_dict = get_item_dict(get_item_pks(rfq_number))
    elif item_id:  # TODO: this should be an else
        item_dict = get_item_dict(create_single_item_dict(item_id, qty_req)) #put a dict
    main_dict = sort_items_in_groups(item_dict)
    category_list = []
    excel_path_list = []
    for key, value in main_dict.items():
        for key1, value1 in value.items():
            if key1 == "EmailCategory" and value1 not in category_list:
                category_list.append(value1)
    for category in category_list:
        filepath = "templates/" + f"RFQ_template_{category}.xlsx"
        excel_path = create_excel(filepath, main_dict, rfq_number, item_id)
        excel_path_list.append(excel_path)
    return excel_path_list


# def send_outlook_email(excel_path, email_list, subject, email_body, other_attachment=[], cc_email=None):
#     """Sends Email using Outlook in the system and attaches the filled excel sheet to the email"""
#     pythoncom.CoInitialize()
#     try:
#         outlook = win32.Dispatch("outlook.application")
#         print("Outlook email creation started")

#         for email in email_list:
#             try:
#                 mail = outlook.CreateItem(0)  # Create a new mail item for each recipient
#                 mail.Subject = subject
#                 mail.Body = email_body
#                 mail.To = email
#                 mail.CC = cc_email

#                 abs_excel_path = os.path.abspath(excel_path)
#                 if os.path.isfile(abs_excel_path):
#                     mail.Attachments.Add(abs_excel_path)
#                 else:
#                     print(f"Attachment not found: {abs_excel_path}")

#                 for attachment in other_attachment:
#                     abs_attachment_path = os.path.abspath(attachment)
#                     if os.path.isfile(abs_attachment_path):
#                         mail.Attachments.Add(abs_attachment_path)
#                     else:
#                         print(f"Additional attachment not found: {abs_attachment_path}")

#                 mail.Send()
#                 print(f"Email sent to {email}")
#                 time.sleep(1)  # Adding a delay to avoid rate limits or other issues
#             except Exception as e:
#                 print(f"An error occurred while sending email to {email}: {e}")
#                 # Retry logic can be added here if needed
#     except Exception as e:
#         print(f"An error occurred: {e}")
#     finally:
#         pythoncom.CoUninitialize()

def send_outlook_email(excel_path, email_list, subject, email_body, other_attachment=[], cc_email=None):
    """Sends a single email using Outlook with all email addresses in BCC and attaches the filled excel sheet."""
    pythoncom.CoInitialize()
    try:
        outlook = win32.Dispatch("outlook.application")
        print("Outlook email creation started")

        mail = outlook.CreateItem(0)  # Create a new mail item
        mail.Subject = subject
        mail.Body = email_body
        mail.BCC = ";".join(email_list)
        mail.To = cc_email # Add all emails to BCC

        # if cc_email:
        #     mail.CC = cc_email

        abs_excel_path = os.path.abspath(excel_path)
        if os.path.isfile(abs_excel_path):
            mail.Attachments.Add(abs_excel_path)
        else:
            print(f"Attachment not found: {abs_excel_path}")

        for attachment in other_attachment:
            abs_attachment_path = os.path.abspath(attachment)
            if os.path.isfile(abs_attachment_path):
                mail.Attachments.Add(abs_attachment_path)
            else:
                print(f"Additional attachment not found: {abs_attachment_path}")

        mail.Send()
        print(f"Email sent to {len(email_list)} recipients.")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        pythoncom.CoUninitialize()

class EmailBodyDialog(simpledialog.Dialog):
    """ Email Body Display window and functions """
    def __init__(self, parent, initial_body):
        self.initial_body = initial_body
        super().__init__(parent, title = "Edit Email Body")

    def body(self, master):
        "Edit Email Body GUI"
        self.title("Edit Email Body")
        self.geometry("600x600")
        self.resizable(True, True)
        tk.Label(master, text="Edit the email body below:").pack(pady=5)
        self.text = ScrolledText(master, wrap=tk.WORD)
        self.text.pack(expand=True, fill=tk.BOTH)
        self.text.insert(tk.END, self.initial_body)

        self.text.bind("<Return>", self.on_enter)
        return self.text

    def apply(self):
        self.result = self.text.get("1.0", tk.END)
    
    def on_enter(self, event):
        """Insert a newline when Enter is pressed"""
        self.text.insert(tk.INSERT, "\n")
        return "break"

def email_body_template(root): #NOTE: In future we can somehow connect this to the database and fill info automatically of the user.
    """Template of the email body"""
    initial_body = ("Dear Supplier, \n\n"
            "Good Day!\n"
            "This email serves as a formal request for a quotation on pricing and lead time for the items listed in the attached Excel spreadsheet.\n\n"
            "For your convenience, we have included several columns in the spreadsheet. While completion of all columns is not mandatory, we kindly request that you fill out those you deem most relevant to accurately reflecting your pricing and lead time for each item.\n"
            "Please note that to ensure a proper evaluation of your offer, a completed spreadsheet is necessary.\n\n"
            "We appreciate your prompt attention to this matter. Should you require any clarification regarding the listed items, please do not hesitate to contact us.\n"
            "\n"
            "Thankyou.\n"
            "<your_name>\n"
            "<your_designation>\n\n"
            "Etezazi Industries Inc.\n"
            "2101 E. 21st St North\n"
            "Wichita, KS 67214\n"
            "(316)-831 9937 Office EXT <your_ext>\n"
            "Email: <your_email>\n"
            "Website: www.etezazi-industries.com\n"
            "AS9100D and ISO9001:2015 and ITAR Registered Company\n")
    dialog = EmailBodyDialog(root, initial_body)
    # root.wait_window(dialog)
    return dialog.result.strip() if dialog.result else initial_body 

def send_mail(rfq_number=None, other_attachment = [], item_id=None, qty_req=None, fin_attachment = []):
    """Main function that sends email for a RFQ Number"""
    email_dict = get_email_groups()    #NOTE: Uncomment this when needed
    # NOTE: below is the test email id's, and you can comment that and uncomment above for real supplier IDS
    # email_dict = {
    #     'mat-al': ['shubham.aggarwal@etezazicorps.com'],
    #     'fin': ['shubham.smvit@gmail.com'],
    #     'hardware': ['shubham.smvit@gmail.com'],
    # }
    if rfq_number:
        subject = f"RFQ - {rfq_number}"
        excel_path_list = create_excel_sheets(rfq_number=rfq_number)
    else:
        subject = f"RFQ - {item_id}"
        excel_path_list = create_excel_sheets(item_id=item_id, qty_req=qty_req)
    root = tk.Tk()
    root.withdraw()
    email_body = email_body_template(root)
    for excel_path in excel_path_list:
        excel_filename1 = os.path.basename(excel_path)
        if rfq_number:
            send_or_not = messagebox.askyesno(f"Send Mail for {excel_filename1}", f"Do you want to send email for {excel_path} ?")
        else:
            send_or_not = True
        if send_or_not:
            response = messagebox.askyesno("View/Edit Excel", f"Do you want to View and Edit the Excel file: {excel_path}")
            if response:
                subprocess.Popen(['start', excel_path], shell=True)
                # messagebox.showinfo("Info", f"Please edit and save the file: {excel_path}")
                while True:
                    edit_complete = messagebox.askyesno("Edit Complete", "Have you finished editing and saving the file?")
                    if edit_complete:
                        break
                    else:
                        messagebox.showinfo("Info", "Please complete your edits and save the file before sending the email.")
            else:
                pass
            excel_filename = os.path.basename(excel_path)
            for key, values in email_dict.items():
                if key in excel_filename:
                    new_subject = f"{subject} ({key})"
                    # email_list_str = "\n".join(values)
                    # new_email_list_str = simpledialog.askstring("Edit Email IDs", f"Current email IDs for {key}:\n{email_list_str}\n\nEdit email IDs (separated by commas):", initialvalue=", ".join(values))
                    # if new_email_list_str:
                    #     new_email_list = [email.strip() for email in new_email_list_str.split(",")]
                    # else:
                    #     new_email_list = values
                    if key == 'fin':
                        attachments = other_attachment + fin_attachment
                    else:
                        attachments = other_attachment
                    new_email_list = get_email_input(root, f"Edit Email IDs for {key}", values)
                    if new_email_list is not None:
                        send_outlook_email(excel_path, new_email_list, new_subject, email_body, other_attachment=attachments, cc_email='quote@etezazicorps.com') #NOTE: add cc_email
                    # send_outlook_email(excel_path, new_email_list, subject, other_attachment=other_attachment)
        else:
            pass
    root.destroy()

def get_email_input(parent, title, initial_emails):
    dialog = EmailDialog(parent, title, initial_emails)
    parent.wait_window(dialog.top)
    return dialog.result

class EmailDialog:
    """Edit Email ID GUI and functions"""
    def __init__(self, parent, title, initial_emails):
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.geometry("400x300")

        self.label = tk.Label(self.top, text="Edit email IDs:")
        self.label.pack(pady=10)

        self.listbox = tk.Listbox(self.top, selectmode=tk.EXTENDED)
        self.listbox.pack(expand=True, fill='both', padx=10, pady=10)
        for email in initial_emails:
            self.listbox.insert(tk.END, email)
        
        self.listbox.bind('<KeyRelease-BackSpace>', self.remove_email)

        self.entry = tk.Entry(self.top)
        self.entry.pack(pady=5)
        self.entry.bind('<Return>', self.add_email)

        self.add_button = ttk.Button(self.top, text="Add Email", command=self.add_email)
        self.add_button.pack(side=tk.LEFT, padx=5, pady=5)

        self.remove_button = ttk.Button(self.top, text="Remove Selected Email", command=self.remove_email)
        self.remove_button.pack(side=tk.LEFT, padx=5, pady=5)

        self.ok_button = ttk.Button(self.top, text="OK", command=self.on_ok)
        self.ok_button.pack(side=tk.RIGHT, padx=5, pady=5)

        self.result = None

    def add_email(self, event=None):
        """Adds email ID"""
        email = self.entry.get().strip()
        if email and email not in self.listbox.get(0, tk.END):
            self.listbox.insert(tk.END, email)
            self.entry.delete(0, tk.END)

    def remove_email(self, event=None):
        """Removes selected email ID"""
        selected_idx = self.listbox.curselection()
        if selected_idx:
            for idx in reversed(selected_idx):
                self.listbox.delete(idx)

    def on_ok(self):
        """Confirms the email id in the box are final and sends them the email"""
        self.result = list(self.listbox.get(0, tk.END))
        self.top.destroy()
        
