from src.general_class import TableManger
import openpyxl
import os
import win32com.client as win32
import pythoncom
import time

def get_item_pks(rfq_pk):
    rfq_line_table = TableManger("RequestForQuoteLine")
    quote_assembly_table = TableManger("QuoteAssembly")
    quote_fk = rfq_line_table.get("QuoteFK", RequestForQuoteFK=rfq_pk)
    all_item_pks = []
    item_fks_dict = {}
    if quote_fk:
        for fk in quote_fk:
            item_pks = quote_assembly_table.get("ItemFK", "QuantityRequired", QuoteFK=fk[0]) 
            print(item_pks)           
            if item_pks:
                all_item_pks.extend(item_pks)
        print(all_item_pks)
        for pk in all_item_pks:
            if pk[0] is not None:
               item_fks_dict[pk[0]] = pk[1]
    print(f"Dict: {item_fks_dict}")
    return item_fks_dict

def get_item_dict(item_fks_dict):
    item_table = TableManger("Item")
    item_dict = {}
    
    for fk, value in item_fks_dict.items():
        details = item_table.get(
            "PartNumber", "Description", "ItemTypeFK", "PartLength",
            "PartWidth", "Thickness", "StockLength", "StockWidth", "Comment",
            ItemPK=fk
        )
        if details:
            columns = ["PartNumber", "Description", "ItemTypeFK", "PartLength",
                       "PartWidth", "Thickness", "StockLength", "StockWidth", "Comment"]
            item_details = details[0]  
            item_dict[fk] = {column: value for column, value in zip(columns, item_details)}
            item_dict[fk]["QuantityRequired"] = value
    
    return item_dict

def get_email_groups():
    """ Returns all emails that belong to the following groups
        'material', 'hardware', 'op', 'tooling'. 
        {'material': <values>, 'op': <values> ...}
     """
    party_buyer_table = TableManger("PartyBuyer") 
    party_table = TableManger("Party")
    parties = {                              # name : partyFK (we get buyerfk from this)
        "fin" : 3755,
        "mat-al": 3744,
        "mat-steel": 3747,
        "ht": 3764,
        'hardware': 3867,
    }
    email_dict = {}
    for key, partyfk in parties.items():
        buyerfk = party_buyer_table.get("BuyerFK", PartyFK=partyfk)
        email_dict[key] = [] 
        for fk in buyerfk:
            email = party_table.get("Email", PartyPK=fk[0])[0][0]
            email_dict[key].append(email)
 
    return email_dict
 
 
# the categories in the below function are meant to be directly used in the email_dict returned in get_email_groups
# the category that is appended to the item dict will match the key in email dict returned in get_email_groups
 

def sort_items_in_groups(item_dict: dict):
    """Identifies item type and then appends the category as a key, value pair to the item_dict."""
    categories = ["standard", "mat", "hardware", "msc", "fin", "kit", "tooling"]
    
    for itempk, column_values in item_dict.items():
        item_type_index = item_dict[itempk]["ItemTypeFK"] - 1  # to get the corresponding index in categories
        category = categories[item_type_index]
        
        part_number_split = item_dict[itempk]["PartNumber"].lower().split()
        email_category = None  # Initialize email_category to avoid reference before assignment

        if category == "fin":
            if "ht" in part_number_split:
                email_category = "ht"
            else:
                email_category = "fin"
        elif category in ["hardware", "tooling"]:
            email_category = "hardware"
        elif category == "mat":
            if "al" in part_number_split:
                email_category = "mat-al"
            elif "steel" in part_number_split or "st" in part_number_split:
                email_category = "mat-steel"
        else:
            email_category = "mat-al"  # Default email_category to the category if not specified
        
        item_dict[itempk]["Category"] = category
        item_dict[itempk]["EmailCategory"] = email_category
 
    return item_dict


def create_excel(filepath, item_dict, rfq_number):
    # Load the existing Excel file
    workbook = openpyxl.load_workbook(filepath)
    new_dir = os.path.join(r"C:\Users\saggarwal\PythonProjects\email_app_new\RFQ_Excel", str(rfq_number))
    new_filepath = os.path.join(new_dir, "_".join(filepath.split("_")[2:]))
    
    # Ensure the directory exists
    os.makedirs(new_dir, exist_ok=True)
    sheet = workbook.active
    sheet.delete_rows(3, sheet.max_row - 2)
    row_idx = 3
    for key, values in item_dict.items():
        for key1, val in values.items():
            if filepath=="RFQ_template_hardware.xlsx":
                if key1=="EmailCategory" and val=="hardware":
                    
                    sheet.cell(row=row_idx, column=1).value = values['PartNumber']
                    sheet.cell(row=row_idx, column=2).value = values['Description']
                    sheet.cell(row=row_idx, column=6).value = values['PartLength']
                    sheet.cell(row=row_idx, column=5).value = values['PartWidth']
                    sheet.cell(row=row_idx, column=4).value = values['Thickness']
                    sheet.cell(row=row_idx, column=7).value = values['Comment']
                    sheet.cell(row=row_idx, column=3).value = values['QuantityRequired']
                    row_idx += 1
            elif filepath=="RFQ_template_fin.xlsx":
                if key1=="EmailCategory" and val=="fin":
                    sheet.cell(row=row_idx, column=1).value = values['PartNumber']
                    sheet.cell(row=row_idx, column=2).value = values['Description']
                    sheet.cell(row=row_idx, column=6).value = values['PartLength']
                    sheet.cell(row=row_idx, column=5).value = values['PartWidth']
                    sheet.cell(row=row_idx, column=4).value = values['Thickness']
                    sheet.cell(row=row_idx, column=7).value = values['Comment']
                    sheet.cell(row=row_idx, column=3).value = values['QuantityRequired']
                    row_idx += 1
            elif filepath=="RFQ_template_mat-al.xlsx":
                if key1=="EmailCategory" and val=="mat-al":
                    sheet.cell(row=row_idx, column=1).value = values['PartNumber']
                    sheet.cell(row=row_idx, column=2).value = values['Description']
                    sheet.cell(row=row_idx, column=6).value = values['StockLength']
                    sheet.cell(row=row_idx, column=5).value = values['StockWidth']
                    sheet.cell(row=row_idx, column=4).value = values['Thickness']
                    sheet.cell(row=row_idx, column=7).value = values['Comment']
                    sheet.cell(row=row_idx, column=3).value = values['QuantityRequired']
                    row_idx += 1
            elif filepath=="RFQ_template_mat-steel.xlsx":
                if key1=="EmailCategory" and val=="mat-steel":
                    sheet.cell(row=row_idx, column=1).value = values['PartNumber']
                    sheet.cell(row=row_idx, column=2).value = values['Description']
                    sheet.cell(row=row_idx, column=6).value = values['StockLength']
                    sheet.cell(row=row_idx, column=5).value = values['StockWidth']
                    sheet.cell(row=row_idx, column=4).value = values['Thickness']
                    sheet.cell(row=row_idx, column=7).value = values['Comment']
                    sheet.cell(row=row_idx, column=3).value = values['QuantityRequired']
                    row_idx += 1
            elif filepath=="RFQ_template_ht.xlsx":
                if key1=="EmailCategory" and val=="ht":
                    sheet.cell(row=row_idx, column=1).value = values['PartNumber']
                    sheet.cell(row=row_idx, column=2).value = values['Description']
                    sheet.cell(row=row_idx, column=6).value = values['PartLength']
                    sheet.cell(row=row_idx, column=5).value = values['PartWidth']
                    sheet.cell(row=row_idx, column=4).value = values['Thickness']
                    sheet.cell(row=row_idx, column=7).value = values['Comment']
                    sheet.cell(row=row_idx, column=3).value = values['QuantityRequired']
                    row_idx += 1

    workbook.save(new_filepath)
    return new_filepath



def send_all_emails(filepath_folder_of_excel_sheets: str) -> None:
    """Sends all emails to the IDs listed in the email sheet for each category"""
    pass

def create_excel_sheets(rfq_number):
    item_dict = get_item_dict(get_item_pks(rfq_number))
    print(item_dict)
    main_dict = sort_items_in_groups(item_dict)
    category_list = []
    excel_path_list = []
    for key, value in main_dict.items():
        for key1, value1 in value.items():
            if key1=="EmailCategory" and value1 not in category_list:
                category_list.append(value1)
    for category in category_list:
        if category == "hardware":
            filepath = "RFQ_template_hardware.xlsx"
            excel_path = create_excel(filepath, main_dict, rfq_number)
            excel_path_list.append(excel_path)
        elif category == "ht":
            filepath = "RFQ_template_ht.xlsx"
            excel_path=create_excel(filepath, main_dict, rfq_number)
            excel_path_list.append(excel_path)
        elif category == "mat-steel":
            filepath = "RFQ_template_mat-steel.xlsx"
            excel_path = create_excel(filepath, main_dict, rfq_number)
            excel_path_list.append(excel_path)
        elif category == "mat-al":
            filepath = "RFQ_template_mat-al.xlsx"
            excel_path = create_excel(filepath, main_dict, rfq_number)
            excel_path_list.append(excel_path)
        elif category== "fin":
            filepath = "RFQ_template_fin.xlsx"
            excel_path=create_excel(filepath, main_dict, rfq_number)
            excel_path_list.append(excel_path)
    return excel_path_list

def send_outlook_email(excel_path, email_list, subject):
    pythoncom.CoInitialize()
    try:
        outlook = win32.Dispatch("outlook.application")
        print("Outlook email creation started")

        for email in email_list:
            try:
                mail = outlook.CreateItem(0)  # Create a new mail item for each recipient
                mail.Subject = subject
                mail.Body = email_body_template()
                mail.To = email
                mail.Attachments.Add(excel_path)
                mail.Send()
                print(f"Email sent to {email}")
                time.sleep(1)  # Adding a delay to avoid rate limits or other issues
            except Exception as e:
                print(f"An error occurred while sending email to {email}: {e}")
                # Retry logic can be added here if needed
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        pythoncom.CoUninitialize()


def email_body_template():
    email_body = ("Hello, \n"
            "Hope you are doing well,\n"
            "Requesting a quote on pricing and lead time for the items listed in the Excel sheet.\n"
            "Please fill out the Excel sheet, or the quote would not be considered.\n"
            "Best,\n"
            "Etezazi Industries\n")
    return email_body 

def send_mail(rfq_number):
    excel_path_list = create_excel_sheets(rfq_number)
    # email_dict = get_email_groups()    #NOTE: Uncomment this when needed
    # NOTE: below is the test email id's, and you can comment that and uncomment above for real supplier IDS
    email_dict = {
        'mat-al': ['shubham.aggarwal@etezazicorps.com', 'yug.banker@etezazicorps.com'],
        'fin': ['yug.banker@etezazicorps.com', 'shubham.smvit@gmail.com'],
        'hardware': ['shubham.smvit@gmail.com'],
    }
    subject = f"RFQ - {rfq_number}"
    
    for excel_path in excel_path_list:
        excel_filename = os.path.basename(excel_path)
        for key, values in email_dict.items():
            if key in excel_filename:
                send_outlook_email(excel_path, values, subject)


if __name__=="__main__":
    send_mail(5995)

 